################################[ Cabeçalho ]##################################

'''
Classe para manipulação de planilhas Excel com openpyxl.
Esta classe permite abrir, salvar e exibir o conteúdo de uma planilha Excel.

Autor: Welton C. O. Rosa
Data: 02/04/2025

Ultima Atualização: 11/08/2025
Atualizado por: Welton C. O. Rosa

Versão: 1.0
Python Version: 3.9.7
Bibliotecas: openpyxl, pandas, selenium, webdriver_manager
Funções disponíveis: criar_planilha, abrir_planilha, salvar_planilha, exibir_conteudo

'''

###############################[ Bibliotecas ]##################################|

# -*- coding: utf-8 -*-
import logging
import os
from openpyxl import load_workbook
from openpyxl import Workbook

#################################[   Log   ]#####################################|

# Configuração do logger
log_file_path = os.path.join(os.getcwd(), 'planilhas.log')  # Caminho absoluto para o arquivo de log
logging.basicConfig(
    filename=log_file_path,                                 # Nome do arquivo de log
    level=logging.INFO,                                     # Nível mínimo de log
    format='%(asctime)s - %(levelname)s - %(message)s'
)

#################################[ Classes ]####################################|

class Planilha:
    def __init__(self, caminho, planilha):
        self.caminho = caminho
        self.planilha = planilha
        self.abrir_planilha()        
    
    def abrir_planilha(self):
        try:
            self.planilha = load_workbook(self.caminho)
        except FileNotFoundError:
            print(f"Arquivo não encontrado: {self.caminho}")
            self.planilha = None
        except Exception as e:
            print(f"Erro ao abrir a planilha: {e}")
            self.planilha = None

    def salvar_planilha(self):
        try:
            self.planilha.save(self.caminho)
            return True
        except Exception as e:
            print(f"Erro ao salvar a planilha: {e}")
            return False
    
    def exibir_conteudo(self):
        if self.planilha is None:
            return  # Se a planilha não foi carregada, encerra a função
        try:
            # Itera sobre todas as planilhas
            for sheet_name in self.planilha.sheetnames:
                print(f"Planilha: {sheet_name}")
                sheet = self.planilha[sheet_name]
                
                # Itera sobre as linhas e colunas da planilha
                for row in sheet.iter_rows(values_only=True):
                    print(row)        
                print("-" * 40)
        except Exception as e:
            print(f"Erro ao listar o conteúdo: {e}")
            return False
        
    def inserir_dados(self, linha, coluna, valor):
        """
        Insere dados em uma célula específica da planilha.

        Parâmetros:
        linha (int): Número da linha onde o dado será inserido.
        coluna (int): Número da coluna onde o dado será inserido.
        valor (str): Valor a ser inserido na célula.
        """
        if self.planilha is None:
            print("Nenhuma planilha carregada. Não é possível inserir dados.")
            return False

        try:
            # Seleciona a planilha ativa
            sheet = self.planilha.active

            # Insere o valor na célula especificada
            sheet.cell(row=linha, column=coluna, value=valor)
            print(f"Valor '{valor}' inserido na célula ({linha}, {coluna}).")
            return True
        except Exception as e:
            print(f"Erro ao inserir dados na célula ({linha}, {coluna}): {e}")
            return False

#################################[ Funções ]####################################|    

def criar_planilha(caminho, nome_planilha):
    """
    Cria um novo arquivo Excel com uma planilha específica.
    
    Parâmetros:
    caminho (str): Caminho completo para salvar o arquivo Excel.
    nome_planilha (str): Nome da planilha a ser criada.
    """
    try:
        # Cria um novo arquivo Excel
        workbook = Workbook()
        
        # Renomeia a planilha padrão para o nome desejado
        sheet = workbook.active
        sheet.title = nome_planilha
        
        # Salva o arquivo no caminho especificado
        workbook.save(caminho)
        print(f"Arquivo Excel criado com a planilha '{nome_planilha}' em: {caminho}")
        return True
    except Exception as e:
        print(f"Erro ao criar a planilha: {e}")
        return False

#################################[ Testes ]####################################|        
""" 
    # Teste da classe Planilha
    if __name__ == "__main__":
        import console

    tela = console.Console("Teste da Classe Planilha")
    tela.limpar_tela()
    tela.exibir_titulo()
   

    # Caminho e nome da planilha
    caminho = r'C:\Users\welton.rosa\Documents\nova_planilha.xlsx'
    nome_planilha = 'Novos Dados'

    # Chama a função para criar a planilha
    criar_planilha(caminho, nome_planilha)

    # Inserir dados na planilha
    planilha = Planilha(caminho, nome_planilha)

    planilha.inserir_dados(1, 1, "Teste de Inserção")
    planilha.inserir_dados(2, 1, "Outro Teste")
    planilha.inserir_dados(3, 1, "Mais um Teste")
    planilha.inserir_dados(4, 1, "Teste Final")

    planilha.salvar_planilha()
    planilha.exibir_conteudo()
 """
